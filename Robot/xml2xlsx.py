import os
from openpyxl import Workbook
import xml.etree.ElementTree as ET


class XML2XLSX():
    def __init__(self, directory, output_file_name):
        self.current_dir = os.path.dirname(__file__)
        self.directory = fr'{self.current_dir}/{directory}'
        self.output_file = fr'{self.directory}/{output_file_name}'

    # Функция для парсинга XML и записи данных в Excel
    def parse_and_write(self):
        # Создаем новый Excel-файл
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'currency'

        # Заголовки
        worksheet.append([
            'Дата USD/RUB', 'Курс USD/RUB', 'Время USD/RUB',
            'Дата JPY/RUB', 'Курс JPY/RUB', 'Время JPY/RUB',
            'USD/RUB к JPY/RUB'
        ])

        # Словари для хранения данных по валютным парам
        usd_data = []
        jpy_data = []

        # Парсинг всех файлов в директории
        for root_dir, dirs, files in os.walk(self.directory):
            for file in files:
                if file.endswith('.xml'):
                    xml_file = os.path.join(root_dir, file)
                    
                    # Определяем валютную пару на основе названия файла
                    if 'USD_RUB' in file:
                        currency_pair = 'USD/RUB'
                    elif 'JPY_RUB' in file:
                        currency_pair = 'JPY/RUB'
                    else:
                        continue  # Пропускаем файлы с неподходящими именами

                    # Парсим XML
                    tree = ET.parse(xml_file)
                    root = tree.getroot()

                    # Извлекаем данные
                    for row in root.findall('.//rows//row'):
                        tradedate = row.get('tradedate')
                        tradetime = row.get('tradetime')
                        rate = row.get('rate')

                        # Преобразуем курс в float
                        rate = float(rate) if rate else None

                        # Сохраняем данные в соответствующий список
                        if currency_pair == 'USD/RUB':
                            usd_data.append((tradedate, rate, tradetime))
                        elif currency_pair == 'JPY/RUB':
                            jpy_data.append((tradedate, rate, tradetime))

        # Объединяем данные по парам
        for usd, jpy in zip(usd_data, jpy_data):
            usd_date, usd_rate, usd_time = usd
            jpy_date, jpy_rate, jpy_time = jpy
            rate_ratio = usd_rate / jpy_rate if usd_rate and jpy_rate else None
            worksheet.append([usd_date, usd_rate, usd_time, jpy_date, jpy_rate, jpy_time, rate_ratio])

        # Форматируем столбцы
        self.format_worksheet(worksheet)

        # Сохраняем Excel
        workbook.save(self.output_file)

    # Функция для форматирования столбцов 
    def format_worksheet(self, worksheet):
        # Автоширина для колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Получаем букву колонки
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  # Добавляем запас
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Форматирование числовых данных
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=worksheet.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "# ##0.000"  # Финансовый формат

    
    def run(self):
        self.parse_and_write()

    
#xml2xlsx = XML2XLSX('downloads', 'output.xlsx')
#xml2xlsx.run()